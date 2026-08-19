"""
Page Streamlit — Analyse automatique d'un Audit Énergétique réglementaire
(rénovation globale CEE) + export de la fiche "Liste des travaux et entreprises".

Extraction 100% par règles (regex), SANS appel IA / sans consommation de tokens.

Fonctionnement :
1. L'utilisateur dépose le PDF de l'audit énergétique.
2. `audit_parser.py` reconstruit le texte par position (x/y) puis applique des
   regex pour extraire bâtiment + scénarios (CEP/CEF avant/après, %, coût, travaux).
3. Rapport complet affiché, avec une table éditable par étape pour corriger les
   champs que le regex n'aurait pas bien identifiés (cas Climawin sur le CEP/CEF
   « après travaux », étiquette-lettre parfois dessinée en vectoriel, etc.).
4. Bouton par scénario -> génère le fichier "Liste des travaux et entreprises.xlsx".

Pré-requis :
- Le fichier gabarit Excel doit être placé dans ./templates/
  Liste_des_travaux_et_entreprises_Modèle.xlsx
"""

import copy
import io
from datetime import date
from pathlib import Path

import streamlit as st

from core.audit_parser import Travail, parse_audit_pdf
from core.pdf_export import generate_pdf as generate_fiche_pdf

TEMPLATE_PATH = (
    Path(__file__).resolve().parent.parent
    / "database" / "templates" / "Liste_des_travaux_et_entreprises_Modèle.xlsx"
)

st.set_page_config(page_title="Audit énergétique — Rénovation globale", layout="wide")
st.title("🏢 Analyse automatique d'audit énergétique — Rénovation globale")
st.caption(
    "Extraction 100% par règles (regex), aucun appel IA. Dépose un audit énergétique "
    "réglementaire (PDF) : l'outil détecte les scénarios de travaux, leurs CEP/CEF "
    "avant/après, les postes de travaux, et permet de générer la fiche Excel "
    "« Liste des travaux et entreprises » pour le scénario choisi."
)

with st.expander("ℹ️ Fiabilité de l'extraction"):
    st.markdown(
        "- Fonctionne sur les formats testés (LICIEL Diagnostics, Climawin) qui suivent "
        "la structure imposée par l'arrêté du 4 mai 2022 (mêmes intitulés de sections).\n"
        "- L'étiquette-lettre (A→G) est parfois un dessin vectoriel non lisible en texte : "
        "elle est récupérée quand une phrase du type *« atteindre la lettre A »* existe, "
        "sinon laissée vide.\n"
        "- Le CEP/CEF « après travaux » peut être imprécis sur certains tableaux multi-lignes "
        "(notamment le format Climawin).\n"
        "- **Toutes les valeurs sont éditables ci-dessous avant génération de l'Excel** — "
        "aucun appel IA n'est fait, la correction se fait à la main si besoin."
    )

# ---------------------------------------------------------------------------
# Upload + extraction
# ---------------------------------------------------------------------------

uploaded = st.file_uploader("Audit énergétique (PDF)", type=["pdf"])

if uploaded is None:
    st.info("Dépose un fichier PDF d'audit énergétique pour lancer l'analyse.")
    st.stop()

if not TEMPLATE_PATH.exists():
    st.warning(
        f"⚠️ Gabarit Excel introuvable : `{TEMPLATE_PATH}`. Place le fichier "
        "« Liste_des_travaux_et_entreprises_Modèle.xlsx » dans `templates/` pour "
        "activer la génération Excel (le rapport reste consultable)."
    )

pdf_bytes = uploaded.read()

with st.spinner("Analyse de l'audit (règles, sans IA)..."):
    batiment, scenarios = parse_audit_pdf(io.BytesIO(pdf_bytes))

if not scenarios:
    st.error(
        "Aucun scénario n'a pu être détecté avec les règles actuelles — la mise en page "
        "de ce document diffère probablement de celles déjà testées (LICIEL / Climawin). "
        "Il faudra ajuster les regex de `audit_parser.py` pour ce format."
    )
    st.stop()

# ---------------------------------------------------------------------------
# Édition du bâtiment
# ---------------------------------------------------------------------------

st.header("📋 Bâtiment")
# Clé de widget dérivée du fichier déposé : force Streamlit à réinitialiser les
# champs (au lieu de garder en mémoire les valeurs d'un fichier précédent) dès
# qu'un nouveau PDF est chargé dans la même session.
fkey = f"{uploaded.name}_{uploaded.size}"
c1, c2 = st.columns(2)
with c1:
    batiment.adresse = st.text_input("Adresse", batiment.adresse or "", key=f"adresse_{fkey}")
    batiment.beneficiaire = st.text_input("Bénéficiaire / Propriétaire", batiment.beneficiaire or "", key=f"benef_{fkey}")
    batiment.surface_m2 = st.number_input("Surface de référence (m²)", value=float(batiment.surface_m2 or 0), step=1.0, key=f"surface_{fkey}")
with c2:
    batiment.cep_initial = st.number_input("CEP initial (kWhEP/m²/an)", value=float(batiment.cep_initial or 0), step=1.0, key=f"cepi_{fkey}")
    batiment.cef_initial = st.number_input("CEF initial (kWhEF/m²/an)", value=float(batiment.cef_initial or 0), step=1.0, key=f"cefi_{fkey}")
    batiment.etiquette_initiale = st.text_input("Étiquette initiale (A-G)", batiment.etiquette_initiale or "", key=f"etqi_{fkey}")

st.divider()

# ---------------------------------------------------------------------------
# Tableau unique : un scénario par colonne (CEP/CEF/étiquette/économie + travaux)
# ---------------------------------------------------------------------------

st.header("📊 Comparatif des scénarios")
st.caption("Une colonne par scénario · ✅ / ❌ pour la présence de chaque type de travaux")

# Mêmes catégories que le "Poste" affiché avant la parenthèse dans "Nature des travaux".
TRAVAUX_TYPES = [
    ("Murs", ["Murs"]),
    ("Plancher bas", ["Plancher", "Planchers bas"]),
    ("Toiture / Combles", ["Toiture / Combles"]),
    ("Menuiseries", ["Menuiseries"]),
    ("Ventilation", ["Ventilation"]),
    ("Chauffage", ["Chauffage"]),
    ("Eau chaude sanitaire", ["Eau chaude sanitaire"]),
]

data = {}
for sc in scenarios:
    derniere = sc.etapes[-1]  # état final du scénario (après toutes ses étapes)
    multi_etapes = len(sc.etapes) > 1
    col_label = sc.nom.replace("\n", " ")
    col = {
        "Étape(s)": " → ".join(e.libelle for e in sc.etapes),
        "CEP avant (kWhEP/m²/an)": batiment.cep_initial,
        "CEP après": derniere.cep_apres,
        "CEF avant (kWhEF/m²/an)": batiment.cef_initial,
        "CEF après": derniere.cef_apres,
        "Étiquette avant": batiment.etiquette_initiale,
        "Étiquette après": derniere.etiquette_apres,
        "Économie": f"{derniere.economie_pct}%" if derniere.economie_pct is not None else "—",
    }
    for label, postes in TRAVAUX_TYPES:
        # étapes (1-indexées) où ce type de travaux apparaît
        etapes_concernees = [
            i for i, e in enumerate(sc.etapes, start=1)
            if any(t.poste in postes for t in e.travaux)
        ]
        if not etapes_concernees:
            col[label] = "❌"
        elif multi_etapes:
            col[label] = "✅ (Étape " + ", ".join(str(i) for i in etapes_concernees) + ")"
        else:
            col[label] = "✅"
    data[col_label] = col

try:
    import pandas as pd

    df = pd.DataFrame(data)
    st.dataframe(df, use_container_width=True)
except ImportError:
    rows = list(next(iter(data.values())).keys())
    st.dataframe(
        [{"": row, **{col: data[col][row] for col in data}} for row in rows],
        use_container_width=True,
        hide_index=True,
    )

st.divider()

# ---------------------------------------------------------------------------
# Détail par scénario, édition, export
# ---------------------------------------------------------------------------

st.header("🔍 Détail, correction et export Excel")

tabs = st.tabs([sc.nom.replace("\n", " ") for sc in scenarios])

for tab, sc in zip(tabs, scenarios):
    with tab:
        for ei, e in enumerate(sc.etapes):
            with st.container(border=True):
                st.markdown(f"**{e.libelle}**")
                m1, m2, m3, m4 = st.columns(4)
                e.cep_apres = m1.number_input(
                    "CEP après", value=float(e.cep_apres or 0), step=1.0, key=f"cep_{fkey}_{sc.id}_{ei}"
                )
                e.cef_apres = m2.number_input(
                    "CEF après", value=float(e.cef_apres or 0), step=1.0, key=f"cef_{fkey}_{sc.id}_{ei}"
                )
                e.etiquette_apres = m3.text_input(
                    "Étiquette après", e.etiquette_apres or "", key=f"lbl_{fkey}_{sc.id}_{ei}"
                )
                e.economie_pct = m4.number_input(
                    "Économie %", value=float(e.economie_pct or 0), step=1.0, key=f"eco_{fkey}_{sc.id}_{ei}"
                )

                st.markdown("_Travaux détectés (modifiable) :_")
                edited = st.data_editor(
                    [
                        {
                            "Poste": t.poste,
                            "Nature (phrase courte)": t.nature_courte,
                            "Caractéristiques / Marque et référence": t.caracteristiques,
                            "Surface / Quantité": t.quantite,
                        }
                        for t in e.travaux
                    ],
                    num_rows="dynamic",
                    use_container_width=True,
                    hide_index=True,
                    key=f"travaux_{fkey}_{sc.id}_{ei}",
                )
                e.travaux = [
                    Travail(
                        poste=row.get("Poste") or "",
                        nature_courte=row.get("Nature (phrase courte)") or "",
                        caracteristiques=row.get("Caractéristiques / Marque et référence") or "",
                        quantite=row.get("Surface / Quantité") or "",
                    )
                    for row in edited
                    if (row.get("Poste") or row.get("Nature (phrase courte)"))
                ]

        st.markdown("")
        colA, colB = st.columns(2)

        with colA:
            if st.button(f"📄 Générer le PDF à envoyer au bailleur", key=f"genpdf_{sc.id}"):
                pdf_bytes = generate_fiche_pdf(batiment, scenarios, scenario_choisi_id=sc.id)
                st.success(
                    "PDF généré : page 1 = récapitulatif des scénarios (scénario "
                    "retenu en évidence), page 2 = fiche des travaux et entreprises "
                    "avec champs remplissables numériquement."
                )
                st.download_button(
                    "⬇️ Télécharger le PDF",
                    data=pdf_bytes,
                    file_name=f"Fiche_travaux_{sc.id}.pdf",
                    mime="application/pdf",
                    key=f"dlpdf_{sc.id}",
                )

        with colB:
            gen_excel = st.button(f"📥 Générer l'Excel pour « {sc.nom.strip()} »", key=f"gen_{sc.id}")

        if gen_excel:
            if not TEMPLATE_PATH.exists():
                st.error("Le gabarit Excel est introuvable, impossible de générer le fichier.")
            else:
                import openpyxl

                def _copy_row_style(ws, src_row, dst_row, min_col, max_col):
                    for col in range(min_col, max_col + 1):
                        s = ws.cell(row=src_row, column=col)
                        d = ws.cell(row=dst_row, column=col)
                        d.font = copy.copy(s.font)
                        d.border = copy.copy(s.border)
                        d.fill = copy.copy(s.fill)
                        d.number_format = s.number_format
                        d.alignment = copy.copy(s.alignment)

                wb = openpyxl.load_workbook(TEMPLATE_PATH)
                ws = wb.active

                ws["C4"] = batiment.beneficiaire or ""
                ws["F4"] = sc.nom.replace("\n", " ")
                ws["C5"] = batiment.adresse or ""
                ws["F5"] = date.today().strftime("%d/%m/%Y")

                all_travaux = []
                for e in sc.etapes:
                    suffix = f" — {e.libelle}" if len(sc.etapes) > 1 else ""
                    for t in e.travaux:
                        all_travaux.append(
                            {
                                "nature": f"{t.nature_affichee}{suffix}",
                                "carac": t.caracteristiques,
                                "quantite": t.quantite,
                            }
                        )

                start_row, avail = 9, 6
                if len(all_travaux) > avail:
                    extra = len(all_travaux) - avail
                    ws.insert_rows(start_row + avail, amount=extra)
                    for i in range(extra):
                        _copy_row_style(ws, start_row, start_row + avail + i, 2, 7)

                for i, t in enumerate(all_travaux):
                    r = start_row + i
                    # Travaux préconisés
                    ws.cell(row=r, column=2, value=t["nature"])
                    ws.cell(row=r, column=3, value=t["carac"])
                    ws.cell(row=r, column=4, value=t["quantite"])
                    # Travaux réalisés : mêmes valeurs par défaut (l'audit ne distingue
                    # pas "préconisé" de "réalisé" — à ajuster manuellement après chantier)
                    ws.cell(row=r, column=5, value=t["nature"])
                    ws.cell(row=r, column=6, value=t["carac"])
                    ws.cell(row=r, column=7, value=t["quantite"])

                buf = io.BytesIO()
                wb.save(buf)
                buf.seek(0)

                st.success("Fichier généré.")
                st.download_button(
                    "⬇️ Télécharger le fichier Excel",
                    data=buf.getvalue(),
                    file_name=f"Liste_travaux_{sc.id}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key=f"dl_{sc.id}",
                )
